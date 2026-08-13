"""Professional multi-sheet Excel workbook for the firewall assessment.

The workbook stores both the compliance results (control-by-control) and the
full assessment configuration into a set of styled sheets:

* Executive Dashboard - title, device metadata, compliance summary, risk rating
* Compliance Results - every evaluated control with status/risk
* Findings - non-compliant findings with remediation guidance
* Configuration sheets - one per assessment section (Health, Policy Analysis,
  Security Services, Routing, VPN, Logging, Administration, Zone Protection,
  Backup, HA)
"""

from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reports.risk_summary import RiskSummary

BRAND_NAVY = "0F172A"
BRAND_ACCENT = "6366F1"
SLATE = "475569"
BORDER_GRAY = "E2E8F0"
LIGHT_BG = "F8FAFC"

STATUS_COLORS = {
    "COMPLIANT": {"text": "16A34A", "fill": "DCFCE7"},
    "NON_COMPLIANT": {"text": "E4002B", "fill": "FEE2E2"},
    "NOT_ASSESSED": {"text": "D97706", "fill": "FEF3C7"},
}

RISK_COLORS = {
    "CRITICAL": "B91C1C",
    "HIGH": "EA580C",
    "MEDIUM": "D97706",
    "LOW": "16A34A",
}

CONFIG_SECTIONS = [
    ("Health", "health_status"),
    ("HA", "ha_configuration"),
    ("Security Services", "security_services"),
    ("Routing", "routing_configuration"),
    ("VPN", "vpn_configuration"),
    ("Logging", "logging_configuration"),
    ("Administration", "administration_configuration"),
    ("Zone Protection", "zone_protection_configuration"),
    ("Backup", "backup_configuration"),
]


def _solid(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _thin_border():
    side = Side(style="thin", color=BORDER_GRAY)
    return Border(left=side, right=side, top=side, bottom=side)


def _style_header_row(ws, row_index, columns):
    for col_index in range(1, columns + 1):
        cell = ws.cell(row=row_index, column=col_index)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = _solid(BRAND_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[row_index].height = 22


def _autofit(ws, min_width=10, max_width=46):
    for column_cells in ws.columns:
        length = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        width = max(min_width, min(max_width, length + 3))
        ws.column_dimensions[letter].width = width


def _append_row(ws, values, start_row=None):
    row_index = start_row or (ws.max_row + 1)
    for col_index, value in enumerate(values, start=1):
        ws.cell(row=row_index, column=col_index, value=value)
    return row_index


def _stringify(value):
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return str(value)
    return value


def _write_kv_table(ws, data, start_row):
    row_index = start_row
    _append_row(ws, ["Metric", "Value"], start_row=row_index)
    _style_header_row(ws, row_index, 2)
    row_index += 1
    for key, value in (data or {}).items():
        _append_row(ws, [key, _stringify(value)], start_row=row_index)
        for col_index in (1, 2):
            cell = ws.cell(row=row_index, column=col_index)
            cell.border = _thin_border()
        ws.cell(row=row_index, column=1).font = Font(bold=True, color=SLATE)
        row_index += 1
    return row_index


class ExcelReport:

    def generate(self, assessment_result, output_file="PaloAlto_Assessment.xlsx"):
        assessment_result = assessment_result or {}

        workbook = Workbook()

        self._build_dashboard(workbook, assessment_result)
        self._build_compliance(workbook, assessment_result)
        self._build_findings(workbook, assessment_result)
        self._build_configuration(workbook, assessment_result)

        workbook.save(output_file)
        return output_file

    # ------------------------------------------------------------------
    # Executive Dashboard
    # ------------------------------------------------------------------

    def _build_dashboard(self, workbook, data):
        ws = workbook.active
        ws.title = "Executive Dashboard"
        ws.sheet_view.showGridLines = False

        inventory = data.get("inventory") or {}
        summary = data.get("summary") or {}

        # Title block
        ws.merge_cells("A1:D1")
        title = ws.cell(row=1, column=1, value="PALO ALTO FIREWALL SECURITY ASSESSMENT")
        title.font = Font(bold=True, color="FFFFFF", size=16)
        title.fill = _solid(BRAND_NAVY)
        title.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 34

        ws.merge_cells("A2:D2")
        subtitle = ws.cell(row=2, column=1, value="LTM Security Platform — Network Security Assessment Report")
        subtitle.font = Font(color=SLATE, size=10, italic=True)
        subtitle.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20

        # Device metadata
        meta = [
            ("Hostname", inventory.get("hostname", "—")),
            ("Model", inventory.get("model", "—")),
            ("Version", inventory.get("version", "—")),
            ("Serial", inventory.get("serial", "—")),
            ("Generated", self._now()),
            ("Source", data.get("_source", "sample")),
        ]
        row = 4
        for label, value in meta:
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = Font(bold=True, color=SLATE)
            label_cell.fill = _solid(LIGHT_BG)
            label_cell.border = _thin_border()
            value_cell = ws.cell(row=row, column=2, value=value)
            value_cell.border = _thin_border()
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1

        # Summary snapshot
        row += 1
        summary_header = ws.cell(row=row, column=1, value="COMPLIANCE SNAPSHOT")
        summary_header.font = Font(bold=True, color="FFFFFF", size=11)
        summary_header.fill = _solid(BRAND_ACCENT)
        summary_header.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

        counts = [
            ("Total Controls", summary.get("total_controls", 0), BRAND_ACCENT),
            ("Compliant", summary.get("compliant", 0), STATUS_COLORS["COMPLIANT"]["text"]),
            ("Non-Compliant", summary.get("non_compliant", 0), STATUS_COLORS["NON_COMPLIANT"]["text"]),
            ("Not Assessed", summary.get("not_assessed", 0), STATUS_COLORS["NOT_ASSESSED"]["text"]),
        ]
        for label, value, color in counts:
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = Font(bold=True, color=SLATE)
            label_cell.border = _thin_border()
            value_cell = ws.cell(row=row, column=2, value=value)
            value_cell.font = Font(bold=True, color=color, size=13)
            value_cell.alignment = Alignment(horizontal="center")
            value_cell.border = _thin_border()
            row += 1

        # Overall risk
        overall_risk = (
            summary.get("overall_risk")
            or RiskSummary.calculate_overall_risk(data.get("assessment", []))
        ).upper()
        risk_color = RISK_COLORS.get(overall_risk, SLATE)

        row += 1
        risk_label = ws.cell(row=row, column=1, value="OVERALL RISK")
        risk_label.font = Font(bold=True, color=SLATE)
        risk_label.fill = _solid(LIGHT_BG)
        risk_label.border = _thin_border()
        risk_value = ws.cell(row=row, column=2, value=overall_risk)
        risk_value.font = Font(bold=True, color="FFFFFF", size=12)
        risk_value.fill = _solid(risk_color)
        risk_value.alignment = Alignment(horizontal="center")
        risk_value.border = _thin_border()
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

        ws.column_dimensions["A"].width = 18
        for letter in ("B", "C", "D"):
            ws.column_dimensions[letter].width = 22

    # ------------------------------------------------------------------
    # Compliance Results (full assessment, control-by-control)
    # ------------------------------------------------------------------

    def _build_compliance(self, workbook, data):
        ws = workbook.create_sheet("Compliance Results")

        headers = ["Control", "Status", "Risk", "Metric", "Observed", "Expected", "Operator"]
        _append_row(ws, headers, start_row=1)
        _style_header_row(ws, 1, len(headers))
        ws.freeze_panes = "A2"

        for result in data.get("assessment", []):
            status = (result.get("status") or "NOT_ASSESSED").upper()
            risk = (result.get("risk") or "MEDIUM").upper()
            row_index = _append_row(
                ws,
                [
                    result.get("control"),
                    status,
                    risk,
                    result.get("metric"),
                    str(result.get("observed", "")),
                    str(result.get("expected", "")),
                    result.get("operator"),
                ],
            )

            status_style = STATUS_COLORS.get(status, {})
            if status_style:
                ws.cell(row=row_index, column=2).font = Font(
                    bold=True, color=status_style["text"]
                )
                ws.cell(row=row_index, column=2).fill = _solid(status_style["fill"])
                ws.cell(row=row_index, column=2).alignment = Alignment(horizontal="center")

            risk_color = RISK_COLORS.get(risk)
            if risk_color:
                ws.cell(row=row_index, column=3).font = Font(bold=True, color=risk_color)
                ws.cell(row=row_index, column=3).alignment = Alignment(horizontal="center")

            for col_index in range(1, len(headers) + 1):
                ws.cell(row=row_index, column=col_index).border = _thin_border()

        ws.auto_filter.ref = ws.dimensions
        _autofit(ws)

    # ------------------------------------------------------------------
    # Findings
    # ------------------------------------------------------------------

    def _build_findings(self, workbook, data):
        ws = workbook.create_sheet("Findings")

        headers = ["Control", "Risk", "Status", "Metric", "Observed", "Expected", "Finding", "Remediation"]
        _append_row(ws, headers, start_row=1)
        _style_header_row(ws, 1, len(headers))
        ws.freeze_panes = "A2"

        for finding in data.get("findings", []):
            risk = (finding.get("risk") or "MEDIUM").upper()
            row_index = _append_row(
                ws,
                [
                    finding.get("control"),
                    risk,
                    finding.get("status"),
                    finding.get("metric"),
                    str(finding.get("observed", "")),
                    str(finding.get("expected", "")),
                    finding.get("finding"),
                    finding.get("remediation"),
                ],
            )
            risk_color = RISK_COLORS.get(risk)
            if risk_color:
                ws.cell(row=row_index, column=2).font = Font(bold=True, color=risk_color)
                ws.cell(row=row_index, column=2).alignment = Alignment(horizontal="center")

            for col_index in range(1, len(headers) + 1):
                cell = ws.cell(row=row_index, column=col_index)
                cell.border = _thin_border()
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.auto_filter.ref = ws.dimensions
        _autofit(ws)

    # ------------------------------------------------------------------
    # Configuration sheets (full assessment sections)
    # ------------------------------------------------------------------

    def _build_configuration(self, workbook, data):
        for sheet_name, data_key in CONFIG_SECTIONS:
            section = data.get(data_key)
            if not isinstance(section, dict):
                continue
            ws = workbook.create_sheet(sheet_name)
            ws.sheet_view.showGridLines = False

            if data_key == "policy_configuration":
                self._write_policy(ws, section)
            else:
                _write_kv_table(ws, section, start_row=1)

            _autofit(ws)

    def _write_policy(self, ws, section):
        scalar_keys = [k for k, v in section.items() if not isinstance(v, list)]
        scalar_data = {k: section[k] for k in scalar_keys}

        row = _write_kv_table(ws, scalar_data, start_row=1)

        rules = section.get("security_rules", [])
        row += 1
        title = ws.cell(row=row, column=1, value="SECURITY RULES")
        title.font = Font(bold=True, color="FFFFFF", size=11)
        title.fill = _solid(BRAND_ACCENT)
        title.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

        rule_headers = ["Name", "Action", "Source", "Destination", "Application", "Service", "Description", "Disabled"]
        _append_row(ws, rule_headers, start_row=row)
        _style_header_row(ws, row, len(rule_headers))
        row += 1

        for rule in rules:
            _append_row(
                ws,
                [
                    rule.get("name"),
                    rule.get("action"),
                    ",".join(rule.get("source", [])),
                    ",".join(rule.get("destination", [])),
                    ",".join(rule.get("application", [])),
                    ",".join(rule.get("service", [])),
                    rule.get("description"),
                    rule.get("disabled"),
                ],
                start_row=row,
            )
            for col_index in range(1, len(rule_headers) + 1):
                ws.cell(row=row, column=col_index).border = _thin_border()
            row += 1

    @staticmethod
    def _now():
        return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
